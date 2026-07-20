from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from django.utils.text import slugify

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class AttendanceReportExportService:
    """Creates CSV, Excel and PDF exports from an attendance queryset."""

    HEADERS = [
        "Student Number",
        "Intern",
        "Batch",
        "Session",
        "Attendance Date",
        "Check-in Time",
        "Check-out Time",
        "Status",
        "Attendance Method",
        "Attendance Location",
        "Recorded By",
    ]

    @classmethod
    def export_csv(cls, records, summary, filters):
        import csv

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="{cls._filename("csv")}"'
        )

        writer = csv.writer(response)
        writer.writerow(["CodeAttend Attendance Report"])
        writer.writerow(["Generated", timezone.localtime().strftime("%d %b %Y %H:%M")])
        writer.writerow(["Filters", cls._filter_description(filters)])
        writer.writerow([])
        writer.writerow(["Total", summary["total"]])
        writer.writerow(["Present", summary["present"]])
        writer.writerow(["Late", summary["late"]])
        writer.writerow(["Absent", summary["absent"]])
        writer.writerow(["Attendance Rate", f'{summary["attendance_rate"]}%'])
        writer.writerow([])
        writer.writerow(cls.HEADERS)

        for record in records.iterator():
            writer.writerow(cls._record_row(record))

        return response

    @classmethod
    def export_excel(cls, records, summary, filters):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Attendance Report"
        worksheet.freeze_panes = "A10"

        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        summary_fill = PatternFill("solid", fgColor="E2F0D9")
        white_font = Font(color="FFFFFF", bold=True, size=16)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        worksheet.merge_cells("A1:K1")
        worksheet["A1"] = "CodeAttend Attendance Report"
        worksheet["A1"].fill = title_fill
        worksheet["A1"].font = white_font
        worksheet["A1"].alignment = Alignment(horizontal="center")
        worksheet.row_dimensions[1].height = 28

        worksheet["A2"] = "Generated"
        worksheet["B2"] = timezone.localtime().strftime("%d %b %Y %H:%M")
        worksheet["A3"] = "Filters"
        worksheet.merge_cells("B3:K3")
        worksheet["B3"] = cls._filter_description(filters)

        summary_items = [
            ("Total", summary["total"]),
            ("Present", summary["present"]),
            ("Late", summary["late"]),
            ("Absent", summary["absent"]),
            ("Attendance Rate", f'{summary["attendance_rate"]}%'),
            (
                "Average Check-in",
                summary["average_check_in"].strftime("%H:%M")
                if summary.get("average_check_in")
                else "-",
            ),
        ]

        for column_index, (label, value) in enumerate(summary_items, start=1):
            label_cell = worksheet.cell(row=5, column=column_index)
            value_cell = worksheet.cell(row=6, column=column_index)
            label_cell.value = label
            value_cell.value = value
            label_cell.font = bold_font
            label_cell.fill = summary_fill
            value_cell.fill = summary_fill
            label_cell.alignment = Alignment(horizontal="center")
            value_cell.alignment = Alignment(horizontal="center")
            label_cell.border = thin_border
            value_cell.border = thin_border

        header_row = 9
        for column_index, header in enumerate(cls.HEADERS, start=1):
            cell = worksheet.cell(row=header_row, column=column_index)
            cell.value = header
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_index, record in enumerate(records.iterator(), start=header_row + 1):
            for column_index, value in enumerate(cls._record_row(record), start=1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")

            worksheet.cell(row=row_index, column=5).number_format = "dd mmm yyyy"
            worksheet.cell(row=row_index, column=6).number_format = "hh:mm:ss"
            worksheet.cell(row=row_index, column=7).number_format = "hh:mm:ss"

        widths = [18, 26, 18, 18, 17, 15, 15, 14, 20, 24, 24]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width

        worksheet.auto_filter.ref = f"A{header_row}:K{max(header_row, worksheet.max_row)}"
        worksheet.sheet_view.showGridLines = False
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.print_title_rows = f"1:{header_row}"

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{cls._filename("xlsx")}"'
        )
        return response

    @classmethod
    def export_pdf(cls, records, summary, filters):
        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=10 * mm,
            leftMargin=10 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
            title="CodeAttend Attendance Report",
            author="CodeAttend",
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            alignment=TA_CENTER,
            fontSize=18,
            leading=22,
            spaceAfter=6,
        )
        small_style = ParagraphStyle(
            "Small",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
        )

        story = [
            Paragraph("CodeAttend Attendance Report", title_style),
            Paragraph(
                f"Generated: {timezone.localtime().strftime('%d %b %Y %H:%M')}",
                small_style,
            ),
            Paragraph(
                f"Filters: {cls._filter_description(filters)}",
                small_style,
            ),
            Spacer(1, 5 * mm),
        ]

        summary_data = [
            ["Total", "Present", "Late", "Absent", "Attendance Rate", "Average Check-in"],
            [
                str(summary["total"]),
                str(summary["present"]),
                str(summary["late"]),
                str(summary["absent"]),
                f'{summary["attendance_rate"]}%',
                summary["average_check_in"].strftime("%H:%M")
                if summary.get("average_check_in")
                else "-",
            ],
        ]
        summary_table = Table(summary_data, colWidths=[43 * mm] * 6)
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C9D6")),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.extend([summary_table, Spacer(1, 5 * mm)])

        table_data = [[Paragraph(header, small_style) for header in cls.HEADERS]]
        for record in records.iterator():
            table_data.append(
                [Paragraph(str(value or "-"), small_style) for value in cls._record_row(record)]
            )

        column_widths = [21, 34, 25, 25, 24, 22, 22, 21, 28, 32, 32]
        report_table = Table(
            table_data,
            repeatRows=1,
            colWidths=[width * mm for width in column_widths],
        )
        report_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C7C7C7")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FB")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(report_table)
        document.build(story)

        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="{cls._filename("pdf")}"'
        )
        return response

    @staticmethod
    def _record_row(record):
        return [
            record.intern.student_number,
            record.intern.user.full_name,
            record.batch.name,
            record.session.name,
            record.attendance_date,
            record.check_in_time,
            record.check_out_time,
            record.get_status_display(),
            record.get_attendance_method_display(),
            record.attendance_location.name if record.attendance_location else "",
            record.recorded_by.full_name if record.recorded_by else "System",
        ]

    @staticmethod
    def _filter_description(filters):
        parts = []
        labels = {
            "search": "Search",
            "intern_label": "Intern",
            "batch_label": "Batch",
            "session_label": "Session",
            "status_label": "Status",
            "attendance_method_label": "Method",
            "start_date": "From",
            "end_date": "To",
        }
        for key, label in labels.items():
            value = filters.get(key)
            if value:
                parts.append(f"{label}: {value}")
        return "; ".join(parts) if parts else "All attendance records"

    @staticmethod
    def _filename(extension):
        date_stamp = timezone.localdate().isoformat()
        base = slugify(f"CodeAttend attendance report {date_stamp}")
        return f"{base}.{extension}"
